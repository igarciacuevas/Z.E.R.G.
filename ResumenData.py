import openpyxl
import os
import sys


def ApuntarResultadosEnExcel(logfiles,problema):
	LibroResultados=openpyxl.Workbook()
	SheetResultados=LibroResultados.active
	SheetResultados.title="Resultados"
    # Cabecera
	contador=1
	SheetResultados.cell(column=contador,row=1,value='NombreArchivo')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='NombreGrafo')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='NumCiudades')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='Tamano_Enjambre')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='MaxIter')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='Crossover')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='Mutacion')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='Config')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='NumIter')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='Estancamiento')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='IterDesdeUltMejora')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='Tiempo')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='Distancia')
	contador+=1
	SheetResultados.cell(column=contador,row=1,value='NumVecinos')
	contador+=1
	
    # fila del excel
	rowcounter=2
    
	for archivo in logfiles:
		print(archivo)
		contadorlinea = 1 # fila del archivo
		BestDist = 20000000.0
		contadorfinal = 0
		contadorultmejora = 0
		ultimaiteracion = 0
		final = False
		listacampos = []
		listacampos.append(archivo)
		listacampos.append(os.path.splitext(archivo)[0]+'.png')
        
		f = open(archivo,'r')
        
		for linea in f:
#			if archivo==r'C:\Users\igarc\ZERG\logs\berlin52\Config1\berlin52_14_21_06_05_12_2018.log':
#				print(linea)
#				print(final)
			if contadorlinea <=10:
				if contadorlinea==6: # Numciudades
					listacampos.append(int(linea.split()[4]))
				elif contadorlinea==7: # Tamaño enjambre
					listacampos.append(int(linea.split()[3]))
				elif contadorlinea==8: # MaxIteraciones
					MaxIter = int(linea.split()[5])
					listacampos.append(MaxIter)
				elif contadorlinea==9: # Crossover
					cross = linea.split()[-1][:-1]
					if cross == "Oder1":
						listacampos.append("Order1")
					else:
						listacampos.append(cross) #Sin el punto final
				elif contadorlinea==10: # Mutacion
					if (linea.split()[0]=="No"):
						listacampos.append("No")
					else:
						mutacion = linea.split()[-1][:-1] #Sin el punto final
						if mutacion=="SingleSwap": #Fix log files
							listacampos.append("RandomSwap")
						else:
							listacampos.append(mutacion) 
				# Correcta clasificacion de config
					if listacampos[-2]=="Cycle":
						if listacampos[-1]=="No":
							listacampos.append("Config4")
						elif listacampos[-1]=="RandomSwap":
							listacampos.append("Config6")
						else:
							listacampos.append("Config5")                        
					elif listacampos[-2]=="PMX": 
						if listacampos[-1]=="No":
							listacampos.append("Config7")
						elif listacampos[-1]=="RandomSwap":
							listacampos.append("Config9")
						else:
							listacampos.append("Config8") 
					else:
						if listacampos[-1]=="No":
							listacampos.append("Config1")
						elif listacampos[-1]=="RandomSwap":
							listacampos.append("Config3")
						else:
							listacampos.append("Config2") 
                            
			# Resto
			elif contadorlinea>=22 and not(final):
#				if contadorlinea==1022:
#					stop=1
				if linea[0:5]==r'FINAL':
					listacampos.append(ultimaiteracion) # NumIter
					listacampos.append(1) # Si Estancamiento
					listacampos.append(contadorultmejora) # IterDesdeUltMejora
					final = True
				else: 
					Iter_Dist = linea.split("--")
                    # Calculo de hace cuantas iteraciones fue la ultima mejora
					if float(Iter_Dist[1])<BestDist:
						contadorultmejora = 0
						BestDist = float(Iter_Dist[1])
					else:
						contadorultmejora += 1
            
					ultimaiteracion = int(Iter_Dist[0])
					# End of busqueda
					if ultimaiteracion == (MaxIter-1):
						final = True
						listacampos.append(ultimaiteracion) # NumIter
						listacampos.append(0) # No Estancamiento
						listacampos.append(contadorultmejora) # IterDesdeUltMejora
			elif final:
				contadorfinal+=1
				if contadorfinal==2: #Tiempo
					listacampos.append(float(linea.split()[7]))
				elif contadorfinal==7: # Distancia
					listacampos.append(float(linea.split()[1]))
				elif contadorfinal==9: # Vecinos
					if "ConVecinos" in listacampos[0]: #Only for ConVecinos
						listacampos.append(float(linea.split()[1]))
				elif linea[0:5]==r'FINAL': # estancamiento == IterMax
					contadorfinal=0
			contadorlinea+=1

		f.close()
		
		# Reclasificacion
		if "SinVecinos" in listacampos[0]:
			if listacampos[5]=="PMX":
				if "Config1" in listacampos[0]:
					listacampos[6]="No"
					listacampos[7]="Config7"
				elif "Config2" in listacampos[0]:
					listacampos[6]="Inversion"
					listacampos[7]="Config8"
				else:
					listacampos[6]="RandomSwap"
					listacampos[7]="Config9"
			elif listacampos[5]=="Order1":
				if "Config1" in listacampos[0]:
					listacampos[6]="No"
					listacampos[7]="Config1"
				elif "Config2" in listacampos[0]:
					listacampos[6]="Inversion"
					listacampos[7]="Config2"
				else:
					listacampos[6]="RandomSwap"
					listacampos[7]="Config3"
			else: #Cycle
				if "Config4" in listacampos[0]:
					listacampos[6]="No"
					listacampos[7]="Config4"
				elif "Config5" in listacampos[0]:
					listacampos[6]="Inversion"
					listacampos[7]="Config5"
				else:
					listacampos[6]="RandomSwap"
					listacampos[7]="Config6"
		else:
			if "Config1" in listacampos[0]:
				listacampos[5]="Order1"
				listacampos[6]="No"
				listacampos[7]="Config1"
			elif "Config2" in listacampos[0]:
				listacampos[5]="Order1"
				listacampos[6]="Inversion"
				listacampos[7]="Config2"
			elif "Config3" in listacampos[0]:
				listacampos[5]="Order1"
				listacampos[6]="RandomSwap"
				listacampos[7]="Config3"
			elif "Config4" in listacampos[0]:
				listacampos[5]="Cycle"
				listacampos[6]="No"
				listacampos[7]="Config4"
			elif "Config5" in listacampos[0]:
				listacampos[5]="Cycle"
				listacampos[6]="Inversion"
				listacampos[7]="Config5"
			elif "Config6" in listacampos[0]:
				listacampos[5]="Cycle"
				listacampos[6]="RandomSwap"
				listacampos[7]="Config6"
			elif "Config7" in listacampos[0]:
				listacampos[5]="PMX"
				listacampos[6]="No"
				listacampos[7]="Config7"
			elif "Config8" in listacampos[0]:
				listacampos[5]="PMX"
				listacampos[6]="Inversion"
				listacampos[7]="Config8"
			elif "Config9" in listacampos[0]:
				listacampos[5]="PMX"
				listacampos[6]="RandomSwap"
				listacampos[7]="Config9"
		# Escribir al excel
		for i in range(len(listacampos)):
			# Write
			SheetResultados.cell(column=i+1,row=rowcounter,value=listacampos[i])
		rowcounter+=1
	
	LibroResultados.save(os.getcwd()+r'\Resultados_' + problema + r'.xlsx')
	return

def ListarArchivos(directorio):
	lista=[]
    
	if not os.path.isdir(directorio):
        # Warning y Exit
		sys.exit('El directorio introducido no existe!')
    
	#OBTENEMOS TODOS LOS DIRECTORIOS Y ARCHIVOS QUE ESTAN EN EL DIRECTORIO INTRODUCIDO
	for dirname, dirnames, filenames in os.walk(directorio):
        # print path to all subdirectories first.
		for subdirname in dirnames:
			lista.append(os.path.join(dirname, subdirname))
    
        # print path to all filenames.
		for filename in filenames:
			lista.append(os.path.join(dirname, filename))
    
	longitudlista=len(lista)
	c=0
    
    # FILTRAMOS LOS PATHS OBTENIDOS 
	while c < longitudlista:
		extension = os.path.splitext(lista[c])[1]
		if not extension=='.log':
			del lista[c]
			longitudlista=longitudlista-1
			c=c-1
            
		c=c+1
        
	return lista  

#rutas = [r'C:\Users\igarc\ZERG\logs\SinVecinos\mundiales']
    
rutas = [r'C:\Users\igarc\ZERG\logs']

for ruta in rutas:
	cachos = ruta.split('\\')
	archivos=ListarArchivos(ruta)
	ApuntarResultadosEnExcel(archivos,cachos[-1])



